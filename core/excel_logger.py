import os
import time
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from core.logger import custom_logger as logger

class ExcelLogger:
    def __init__(self, output_dir="logs"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        
    def _get_daily_filepath(self):
        date_str = time.strftime("%Y-%m-%d")
        return os.path.join(self.output_dir, f"trading_report_{date_str}.xlsx")

    def log_trade(self, trade_data: dict, portfolio_summary: dict):
        """Appends trade transaction and current portfolio state into formatted Excel using pure openpyxl."""
        file_path = self._get_daily_filepath()
        
        headers = [
            "Time", "Symbol", "Instrument", "Action", "Qty", 
            "Entry Index", "Exit Index", "Entry Premium", "Exit Premium", 
            "P&L", "Peak Run (Pts)", "Reason", "Account Balance"
        ]
        
        # Convert timestamp to Indian Standard Time (IST: UTC+5:30)
        ist_tz = timezone(timedelta(hours=5, minutes=30))
        trade_time_ist = datetime.fromtimestamp(trade_data.get("timestamp", time.time()), ist_tz)
        
        trade_values = [
            trade_time_ist.strftime("%H:%M:%S"),
            trade_data.get("symbol"),
            trade_data.get("instrument", "N/A"),
            trade_data.get("action"),
            trade_data.get("qty"),
            round(trade_data.get("entry_price", 0.0), 2),
            round(trade_data.get("exit_price", 0.0), 2),
            round(trade_data.get("entry_premium", 0.0), 2),
            round(trade_data.get("exit_premium", 0.0), 2),
            round(trade_data.get("pnl", 0.0), 2),
            round(trade_data.get("peak_run", 0.0), 2),
            trade_data.get("reason", "STRATEGY EXIT"),
            round(portfolio_summary.get("current_balance", 0.0), 2)
        ]
        
        try:
            if os.path.exists(file_path):
                wb = openpyxl.load_workbook(file_path)
                if "Trades" in wb.sheetnames:
                    ws = wb["Trades"]
                else:
                    ws = wb.create_sheet("Trades")
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Trades"
                # Write header if new file
                ws.append(headers)
                
            # Append new row
            ws.append(trade_values)
            
            # Format and Save Workbook
            self._format_excel_sheet(ws)
            wb.save(file_path)
            wb.close()
            logger.debug(f"Excel report updated and stylized at: {file_path}")
            
        except Exception as e:
            logger.error(f"Error logging trade to Excel: {e}")

    def _format_excel_sheet(self, ws):
        """Applies headers styling, grid lines, alignment, and P&L green/red color styling."""
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10)
        
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
        red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Red
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Format Headers (Row 1)
        for col in range(1, 14):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Format Data rows
        for row in range(2, ws.max_row + 1):
            pnl_cell = ws.cell(row=row, column=10) # Column J (10th column) is P&L
            pnl_val = 0.0
            try:
                pnl_val = float(pnl_cell.value or 0.0)
            except (ValueError, TypeError):
                pass
            
            # Dynamic conditional formatting
            fill_color = green_fill if pnl_val > 0 else (red_fill if pnl_val < 0 else None)
            
            for col in range(1, 14):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                # Alignment & Number Formatting
                if col in [1, 2, 3, 4]: # Time, Symbol, Instrument, Action
                    cell.alignment = Alignment(horizontal="center")
                elif col in [5, 6, 7, 8, 9, 10, 11, 13]: # Qty, Entry/Exit Index, Entry/Exit Premium, P&L, Peak Run, Balance
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
                else: # Reason
                    cell.alignment = Alignment(horizontal="left")
                
                # Apply background color to P&L cell
                if col == 10 and fill_color:
                    cell.fill = fill_color
        
        # Auto-adjust column widths for premium layout
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
